# -*- coding: utf-8 -*-
from threading import Thread
import openpyxl
from openpyxl import load_workbook
import telebot 
import random
import time
import datetime 

from telebot import types
global osn_base
global all_base
global con_base
global test_day
global comm_admin

def main():
    test_day=7
    admin=[405027580,741710024]
    comm_admin=str("ВКЛЮЧИТЬ БОТА:\non (например, on)\n\n"+
               "ВЫКЛЮЧИТЬ БОТА:\noff(например, off)\n\n"+
               "ЗАБАНИТЬ ЧЕЛОВЕКА, КОТОРОМУ ПРЕНАДЛЕЖИТ ДАННАЯ АНКЕТА:\nban_on;*номер анкеты*(например, ban_on;3.1)\n\n"+
               "УБРАТЬ ИЗ БАНА ЧЕЛОВЕКА, КОТОРОМУ ПРЕНАДЛЕЖИТ ДАННАЯ АНКЕТА:\nban_off;*НОМЕР АНКЕТЫ*(например, ban_off;3.1)\n\n"+
               "НАПИСАТЬ ПРИЧИНУ ОТКАЗА ДЛЯ ДАННОЙ АНКЕТЫ:\n*номер анкеты*;*причина отказа*(например, 3.1;ПРИЧИНА)\n\n"+
               "ДОБАВИТЬ НЕКОТОРОЕ КОЛИЧЕСТВО МЕСЯЦЕВ ДЛЯ ОПРЕДЕЛЕННОЙ АНКЕТЫ\n*номер анкеты*+*количество месяцев*(например, 3.1+12)\n\n"+
               "УМЕНЬШИТЬ НЕКОТОРОЕ КОЛИЧЕСТВО МЕСЯЦЕВ ДЛЯ ОПРЕДЕЛЕННОЙ АНКЕТЫ\n*номер анкеты*-*количество месяцев*(например, 3.1+12)\n\n")
    #-----------------------------------------------------------------------------------------------------------
    """
    osn_base='C:/Users/nikita/Desktop/database.xlsx'
    all_base='C:/Users/nikita/Desktop/all.xlsx'
    con_base='C:/Users/nikita/Desktop/contact.xlsx'
    """
    osn_base='lb2/database.xlsx'
    all_base='lb2/all.xlsx'
    con_base='lb2/contact.xlsx'
    
    
    bot=telebot.TeleBot('1292714271:AAFto5D4qOOmTbRDfYVY28DQguWr3FJWKlc')   #бот для принятия анкет и для отправки других
    bot_checker = telebot.TeleBot('1147234538:AAHFUcJE44cGiFFBISV5YCtK8TggG2Jf9ps') #бот для проверки анкет
    
    keyboard_start = telebot.types.ReplyKeyboardMarkup(True, True)    #клавиатура для определения типа покупателя
    keyboard_start.row("Заполнить анкету")
    keyboard_start.row("Проверить статус анкеты")
    keyboard_start.row("Реферальная программа")
    
    keyboard_anketa = telebot.types.ReplyKeyboardMarkup(True, True)    #клавиатура для определения типа покупателя
    keyboard_anketa.row("Рекламодатель")
    keyboard_anketa.row("Покупатель")
    
    
    keyboard_network_type = telebot.types.ReplyKeyboardMarkup(True, True)
    keyboard_network_type.row("Instagram") #кнопки для определения площадки рекламы
    keyboard_network_type.row("Telegram")
    keyboard_network_type.row("YouTube")
    keyboard_network_type.row("Tik-Tok")
    keyboard_network_type.row("Вконтакте")
    
    keyboard_category_ad = telebot.types.ReplyKeyboardMarkup(True, True)
    keyboard_category_ad.row("Личный блог")
    keyboard_category_ad.row("Блог")                 #клавиатура для определения категории рекламы
    keyboard_category_ad.row("Магазин/ресторан")
    keyboard_category_ad.row("Группа")
    keyboard_category_ad.row("Остальное")
    
    
    
    
    
    #-----------------------------------------------------------------------------------------------------------Слушающие функции
    
    @bot.message_handler(content_types=['text'])
    def get_start_message(message):
        if test_on_off()==0:
            bot.send_message(message.from_user.id,"Бот временно отключен (Великие Администраторы делают его еще лучше)\nПопробуйте зайти сюда чуть позднее")
            if message.from_user.id in admin:
                bot.send_message(message.from_user.id,"Но ты админ, поэтому можешь работать))))")
        if test_on_off()==1 or message.from_user.id in admin:
            if message.text == '/info':
                markup = types.InlineKeyboardMarkup()
                btn_my_site= types.InlineKeyboardButton(text='Соглашение', url='https://telegra.ph/Soglashenie-05-27')
                markup.add(btn_my_site)
                bot.send_message(message.chat.id, "Нажми на кнопку и прочитай соглашение.", reply_markup = markup)
            if check_ban(message.from_user.id):   
                
                if message.text == '/start':  #начало диалога с клиентом
                    bot.send_message(message.from_user.id, "Здравствуйте! Добро пожаловать на биржу рекламы. Начиная работу, вы соглашаетесь на правила использования бота, которые можно прочитать, нажав /info\n\nНа бирже вы можете продать и найти рекламу для ваших аккаунтов в социальных сетях! Выберите, что именно Вас интересует, нажав на одну из кнопок",reply_markup=keyboard_start)
                    
                if message.text=='/search':
                    if test_in_main(message.from_user.id,osn_base):
                        bot.send_message(message.from_user.id,"Введите ID Вашей анкеты")
                        bot.register_next_step_handler(message, podbor)
                    else:
                        bot.send_message(message.from_user.id,"Ваша заявка либо не найдена, либо еще не подтверждена администратором")
                if message.text=="Заполнить анкету":
                    bot.send_message(message.from_user.id,"Для начала заполнения анкеты выберите пожалуйста, кто вы\nРекламодатель-можете предоставить место для контента\nПокупатель-хотите купить рекламу",reply_markup=keyboard_anketa)
                    bot.register_next_step_handler(message, get_type_message)
                    
                if message.text=="Проверить статус анкеты":
                    bot.send_message(message.from_user.id,"Введите пожалуйста номер анкеты, которая Вас интересует")
                    bot.register_next_step_handler(message, status)
    
                if message.text=="Реферальная программа":
                    if message.from_user.id not in admin:
                        bot.send_message(message.from_user.id,"Данный раздел пока недоступен")   
                    else:
                        bot.send_message(message.from_user.id,"Данный раздел пока недоступен")   
                
            else:
                if test_on_off()==1:
                    bot.send_message(message.from_user.id,"Вы забанены. Обратитесь в поддержку ( @metand_2 )")
                
                
                
    #обработка кнопок для подтверждения корректности анкеты 
    @bot.callback_query_handler(func=lambda call: True)
    def correct(call):
        lis=call.data.split(" ")
        if lis[0] == 'yes' and len(lis)==2:
            
    
            anketa=lis[1]
            wb = openpyxl.load_workbook(filename = all_base)
            if anketa.split(".")[0]=="1":
                sheet=wb["1"]    
                
                stroka=get_stroka_po_ankete(anketa,all_base)
                network_type_buyer = sheet["B"+stroka].value
                name_buyer = sheet["C"+stroka].value
                subscribers_buyer = sheet["D"+stroka].value
                statystics_buyer = sheet["E"+stroka].value
                info_buyer = sheet["F"+stroka].value
                category_ad_buyer = sheet["G"+stroka].value        
                price_buyer=sheet["H"+stroka].value   
                
                s="📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_buyer + '\n' + "👨‍💻 2.Никнейм: " + name_buyer + '\n'  "👥 3."+ subscribers_buyer + "подписчиков"+ '\n' + "📊 4.Статистика профиля: " + statystics_buyer + '\n' + "🗺 5.Инфомация об аудитории: " + info_buyer + '\n' + "📄 6.Категория рекламы: " + category_ad_buyer + '\n' + "💰 7. Максимальная стоимость рекламы: "+ price_buyer
            else:
                sheet=wb["0"] 
                
                stroka=get_stroka_po_ankete(anketa,all_base)
            
                type_=int(sheet["L"+stroka].value)
                network_type_advertiser=sheet["B"+stroka].value
                name_advertiser=sheet["C"+stroka].value
                subscribers_advertiser=sheet["D"+stroka].value
                statystics_advertiser=sheet["E"+stroka].value
                info_advertiser=sheet["F"+stroka].value
                contacts_advertiser=sheet["G"+stroka].value
                self_category_ad=sheet["H"+stroka].value
                category_ad_advertiser=sheet["I"+stroka].value
                price_advertiser=sheet["J"+stroka].value
                barter_advertiser=sheet["K"+stroka].value
                #print(price_advertiser)
                if type_==3:
                    qwe=price_advertiser.split(",")
                if int(barter_advertiser)!=0:
                    if type_ == 3:
                        s="📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'+ "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' + "💰 9. Стоимость рекламного поста: "+ qwe[0] + '\n' + "💰 Стоимость сторис: " +qwe[1]+'\n' + "💰 Стоимость вечного поста: " + qwe[2] + '\n' + "♻️ 10. Возможность бартера при разнице подписчиков: "+str(barter_advertiser)  
                    elif type_ == 1:
                        s= "📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'+ "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' + "💰 9. Стоимость рекламного поста: "+ price_advertiser + '\n' + "♻️ 10. Возможность бартера при разнице подписчиков: "+  str(barter_advertiser)
                else:
                    if type_ == 3:
                        s= "📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'+ "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' + "💰 9. Стоимость рекламного поста: "+ qwe[0] + '\n' + "💰 Стоимость сторис: " +qwe[1]+'\n' + "💰 Стоимость вечного поста: " + qwe[2] 
                    elif type_ == 1:
                        s= "📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'+ "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' + "💰 9. Стоимость рекламного поста: "+ price_advertiser 
                    
                
                
                
            bot.send_message(get_id_po_ankete(lis[1],all_base),"Ваша анкета отправлена администраторам на проверку. Ожидайте...")
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=s)
    
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb[str(lis[1].split(".")[0])]    
            
    
            sheet["N"+get_stroka_po_ankete(lis[1],all_base)].value="1"
            wb.save(all_base)   
            
            #print("yes")
            sending_application(lis[1])
        elif lis[0]=="no" and len(lis)==2:
            bot.edit_message_reply_markup(chat_id=lis[3], message_id = lis[2], reply_markup = '')
            
            #print("no")
            #print(lis)
            network_type=int(lis[1].split(".")[0])
            #print(network_type)
            mes1 = bot.send_message(get_id_po_ankete(lis[1],all_base), "Введите номер пункта, который Вы хотите изменить.")
            if network_type == 0: 
                bot.register_next_step_handler(mes1, number_edit_advertiser)
            elif network_type == 1: 
                bot.register_next_step_handler(mes1, number_edit_buyer)
                # #################################
        elif len(call.data.split(" "))==3:
            ll=call.data.split(" ")
            wb = openpyxl.load_workbook(filename = osn_base)
            sheet=wb['1']
            anketa=ll[2]        
            id_send=get_id_po_ankete(anketa,osn_base)
            st=str(get_stroka_po_ankete(ll[2],osn_base))
            if ll[0]=="ok":
                if sheet[str("L"+st)].value=="time" and sheet[str("M"+st)].value=="time":
                    bot.send_message(id_send, "Ваш ответ отправлен рекламодателю. Ожидайте ответа. Если он согласится, вам придет сообщение с его контактными данными("+ll[1]+")")
                    
                    sheet[str("L"+st)].value=None
                    sheet[str("M"+st)].value=None
                    wb.save(osn_base)                
                    send_to_reklam(anketa,ll[1])
                    
                elif sheet[str("L"+st)].value==None or sheet[str("M"+st)].value==None or sheet[str("L"+st)].value=="" or sheet[str("M"+st)].value=="":
                    bot.send_message(id_send, "Упс! Вы нажали не на ту клавишу!")
                else:
                    bot.send_message(id_send, "Ваш ответ отправлен рекламодателю. Ожидайте ответа. Если он согласится, вам придет сообщение с его контактными данными("+ll[1]+")")
                    sheet[str("L"+st)].value=None
                    sheet[str("M"+st)].value=None
                    wb.save(osn_base)
                    send_to_reklam(anketa,ll[1])
                    
                
            else:
                if sheet[str("L"+st)].value=="time" and sheet[str("M"+st)].value=="time":
                    bot.send_message(id_send,"Отказ принят! Для повтора поиска можете нажать /search ")
                    sheet[str("L"+st)].value=None
                    sheet[str("M"+st)].value=None
                    wb.save(osn_base)
                else:
                    list_ank=sheet[str("L"+st)].value
                    number_ank_now=sheet[str("M"+st)].value
                    if list_ank==None or number_ank_now==None or list_ank=="" or number_ank_now=="":
                        bot.send_message(id_send, "Упс! Вы нажали не на ту клавишу!")
                    else:
                        sheet[str("M"+st)].value=int(sheet[str("M"+st)].value)+1
                        wb.save(osn_base)
        
                        send_anketa(id_send,anketa,list_ank,number_ank_now+1)  #not next!
                    
                    
                    
        elif len(call.data.split(" "))==4:
            lq=call.data.split(" ") #callback_data= 'no '+reklam_ank+" "+anketa+" 1"
            s=contact(lq[1])
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=s)        
            if lq[0]=="ok":
                bot.send_message(get_id_po_ankete(lq[2],osn_base),"Рекламодатель(ID="+lq[1]+") согласился сотрудничать с вами\nЕго контактные данные:\n"+s)
                bot.send_message(get_id_po_ankete(lq[1],osn_base),"Ваши контактые данные отправлены покупателю")
            else:
                bot.send_message(get_id_po_ankete(lq[2],osn_base),"Рекламодатель(ID="+lq[1]+") не согласился сотрудничать с вами")
            # #####################

    @bot_checker.callback_query_handler(func=lambda call: True)
    def acception(call):
        l_tt=call.data.split(" ")
        
        if l_tt[0] == 'accept':
            anketa=l_tt[1]
            a=get_id_po_ankete(anketa,all_base)
            tt=anketa.split('.')[0]
            obrabotka(anketa)
            #-------------------------добавить анкету в contact
            dob_anketa_contact(anketa)
            #--------------------------------------------------
            bot_checker.send_message(405027580,"Анкета №"+l_tt[1]+" добавлена в Базу Данных!")
            bot_checker.send_message(741710024,"Анкета №"+l_tt[1]+" добавлена в Базу Данных!")
            b=get_stroka_po_ankete(anketa,osn_base)
            if tt=="0":
                
                wb = openpyxl.load_workbook(filename = osn_base)
                sheet = wb['0']    
                sheet["T"+str(b)].value=time.time()+test_day*86400
                wb.save(osn_base)
                
                bot.send_message(a, "Ваша анкета №"+l_tt[1]+ " одобрена!"+"\n"+"Ожидайте!")
                bot.send_message(a, "Запущен тестовый период для вашей анкеты ("+str(test_day)+" дней). Анкета активна до "+sec_to_date(time.time(),test_day*86400))
                news_for_pokup(l_tt[1])
            else:
    
                wb = openpyxl.load_workbook(filename = osn_base)
                sheet = wb['1']    
                sheet["T"+str(b)].value=time.time()+test_day*86400
                wb.save(osn_base)
                            
                bot.send_message(a, "Ваша анкета №"+l_tt[1]+ " одобрена!"+"\n"+"Введите /search , чтобы начать поиск рекламодателей.")
                bot.send_message(a, "Запущен тестовый период для вашей анкеты ("+str(test_day)+" дней). Анкета активна до "+sec_to_date(time.time(),test_day*86400))
                
    
        elif l_tt[0] == 'refusal':
            #print("refusal")
            mes=bot_checker.send_message(741710024, "Укажите причину отказа анкете №"+l_tt[1]+"(например '1.3;Проблема' )")
            bot_checker.send_message(405027580, "Укажите причину отказа анкете №"+l_tt[1]+"(например '1.3;Проблема' )")
            bot_checker.register_next_step_handler(mes, cause_refusal)
    @bot_checker.message_handler(content_types=['text'])
    def ratata(message):
        l=message.text.split(";")
        
        if check_float(l[0]):
            cause=l
            if test_anketa(str(cause[0])):
                bot_checker.send_message(741710024, "Причина отказа анкете №"+str(cause[0])+":"+str(cause[1]))
                bot_checker.send_message(405027580, "Причина отказа анкете №"+str(cause[0])+":"+str(cause[1]))   
                bot.send_message(get_id_po_ankete(str(cause[0]),all_base), "Ваша анкета не принята. \nПричина: " + cause[1] + '\n' + "Вы можете снова заполнить анкету. Для этого введите: /start")
            else:
                bot_checker.send_message(message.from_user.id, "Такой анкеты нет")
                
        else:
            if l[0]=="ban_on":
                
                if check_ban(get_id_po_ankete(l[1],all_base)):
                    add_to_ban(get_id_po_ankete(l[1],all_base))
                    bot_checker.send_message(message.from_user.id, "Данный человек добавлен в БАН успешно")
                else:
                    bot_checker.send_message(message.from_user.id, "Данный человек уже добавлен в БАН")
            elif l[0]=="ban_off":
                if check_ban(get_id_po_ankete(l[1],all_base)):
                    bot_checker.send_message(message.from_user.id, "Данный человек НЕ добавлен в БАН")
                else:
                    delete_from_ban(get_id_po_ankete(l[1],all_base))
                    bot_checker.send_message(message.from_user.id, "Данный человек удален из БАНа")
            elif l[0].lower()=="on": on(message)
            elif l[0].lower()=="off":off(message)
            elif l[0]=="/info":
                bot_checker.send_message(message.from_user.id, comm_admin)
            elif len(message.text.split(" "))==2 and message.text.split(" ")[0].upper()=="INFO":
                anketa=message.text.split(" ")[1]
                s=get_info_po_ankete(anketa,osn_base)
                if test_int(s)==False:
                    bot.send_message(message.from_user.id,s)
                else:
                    bot.send_message(message.from_user.id,"Данная анкета не найдена")
                
            elif plus_check(l[0]):
                l1=l[0].split("+")
                plus(l1[0],l1[1])
                wb = openpyxl.load_workbook(filename = osn_base)
                sheet = wb[l1[0].split(".")[0]]     
                tt=sheet["U"+str(get_stroka_po_ankete(l1[0],osn_base))].value
                
                bot_checker.send_message(405027580, "Время использования биржи для анкеты "+str(l1[0])+" увеличено на "+str(l1[1])+" месяц(-а,-ев)\nАктивна до "+str(tt))
                bot_checker.send_message(741710024, "Время использования биржи для анкеты "+str(l1[0])+" увеличено на "+str(l1[1])+" месяц(-а,-ев)\nАктивна до "+str(tt))
                
            elif minus_check(l[0]):
                l1=l[0].split("-")
                minus(l1[0],l1[1])            
                wb = openpyxl.load_workbook(filename = osn_base)
                sheet = wb[l1[0].split(".")[0]]     
                tt=sheet["U"+str(get_stroka_po_ankete(l1[0],osn_base))].value
                
                bot_checker.send_message(405027580, "Время использования биржи для анкеты "+str(l1[0])+" уменьшено на "+str(l1[1])+" месяц(-а,-ев)\nАнктивна до "+str(tt))
                bot_checker.send_message(741710024, "Время использования биржи для анкеты "+str(l1[0])+" уменьшено на "+str(l1[1])+" месяц(-а,-ев)\nАктивна до "+str(tt))            
    #-----------------------------------------------------------------------------------------------------------Вспомогательные функции
    
    def on(message):
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        ans=str(sheet["X8"].value)
        if ans=="1":
            bot_checker.send_message(message.from_user.id, "Бот уже включен")
        elif ans=="0":
            bot_checker.send_message(405027580, "Бот включен")
            bot_checker.send_message(741710024, "Бот включен")
            sheet["X8"].value="1"
            wb.save(all_base)
    def off(message):
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        ans=str(sheet["X8"].value)    
        if ans=="0":
            bot_checker.send_message(message.from_user.id, "Бот уже выключен")
        elif ans=="1":
            bot_checker.send_message(405027580, "Бот выключен")
            bot_checker.send_message(741710024, "Бот выключен")
            sheet["X8"].value="0"
            wb.save(all_base)
    def test_on_off():
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        if str(sheet["X8"].value)=="1":
            return 1
        else:
            return 0
        
    def check_float(a):
        try:
            a=float(a)
        except:
            return False
        return True
    def check_ban(id_):
        #id_=get_id_po_ankete(anketa,all_base)
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        l=str(sheet["X2"].value).split(",")
        if str(id_) in l:
            return False
        return True
    def delete_from_ban(id_):
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        l=str(sheet["X2"].value).split(",")
        t=-1
        for i in range(len(l)):
            if l[i]==str(id_):
                t=i
        if t!=-1:
            l.pop(t)
        s=""
        for i in range(len(l)-1):
            s=str(s+str(l[i])+",")
        s=str(s+str(l[-1]))
        sheet["X2"].value=str(s)
        wb.save(all_base)
    
    def add_to_ban(id_):
        wb = openpyxl.load_workbook(filename = all_base)
        sheet = wb['0']    
        l=str(sheet["X2"].value).split(",")
        
        l.append(str(id_))
        s=""
        for i in range(len(l)-1):
            s=str(s+str(l[i])+",")
        s=str(s+str(l[-1]))
        sheet["X2"].value=str(s)
        wb.save(all_base)
        
    def news_for_pokup(anketa): # вводится анкета рекламодателя
        inf=get_info_po_ankete(anketa,osn_base)
        res=[]
        reklam=[inf[0],inf[1],inf[8],inf[9],inf[11]]
        #['741710024', 'Telegram', 'Q', 'W', 'E', 'R', 'T', 'Остальное', 'Остальное', '10', '0', '0.20']
        
        wb = openpyxl.load_workbook(filename = osn_base)
        sheet = wb['1']    
        a=int(sheet["Z1"].value)
        #print("test0 "+str(a))
        for i in range(2,a):
            if str(sheet["N"+str(i)].value)=="1":
                #print(str(sheet["B"+str(i)].value)+reklam[1])
                if str(sheet["B"+str(i)].value)==str(reklam[1]):
                    #print("test1")
                    #print(sheet["I"+str(i)].value,reklam[2])
                    if str(sheet["G"+str(i)].value)==str(reklam[2]):
                        #print("test2")
                        
                        if check_price(str(sheet["H"+str(i)].value),reklam[3]):
                            if test_date(sheet["Q"+str(i)].value):
                                keyboard_ask = telebot.types.InlineKeyboardMarkup()
                                button_accept = telebot.types.InlineKeyboardButton(text ="Согласен✅", callback_data = 'ok '+inf[-1]+" "+str(sheet["Q"+str(i)].value))
                                button_refusal = telebot.types.InlineKeyboardButton(text ="Не подходит❌", callback_data= 'no '+inf[-1]+" "+str(sheet["Q"+str(i)].value))
                                keyboard_ask.add(button_accept)
                                keyboard_ask.add(button_refusal)
                                st=st=str(get_stroka_po_ankete(str(sheet["Q"+str(i)].value),osn_base))
                                #bot.send_message(id_send,from_list_to_str_1(get_info_po_ankete(list_ank[number_ank_now],osn_base)),reply_markup=keyboard_ask)
                                sheet[str("L"+st)].value="time"
                                sheet[str("M"+st)].value="time"
                                wb.save(osn_base)
                                bot.send_message(int(sheet["A"+str(i)].value),"Внимание! Для Вас(анкета №"+str(sheet["Q"+str(i)].value)+") найден новый рекламодатель!\n\n"+from_list_to_str_1(get_info_po_ankete(reklam[-1],osn_base)),reply_markup=keyboard_ask)
    
    
    
    def list_to_str_vivod(l):
        s=""
        s=str(s+"Социальная сеть: "+str(l[1])+"\n")
        s=str(s+"Никнейм: "+str(l[2])+"\n")
        s=str(s+"Количество подписчиков: "+str(l[3])+"\n")
        s=str(s+"Статистика профиля: "+str(l[4])+"\n")
        s=str(s+"Инфомация об аудитории: " +str(l[5])+"\n")
        s=str(s+"Категория рекламы: "+str(l[6])+"\n")
        s=str(s+"Максимальная стоимость рекламы: "+str(l[7])+"\n")
        
        return s
    
    def send_to_reklam(anketa,reklam_ank):#Кто согласился(покупатель) ; на какую анкету(рекламодатель)
        
        keyboard_ask = telebot.types.InlineKeyboardMarkup()
        button_accept = telebot.types.InlineKeyboardButton(text ="Подходит✅", callback_data = 'ok '+reklam_ank+" "+anketa+" 1")
        button_refusal = telebot.types.InlineKeyboardButton(text ="Не подходит❌", callback_data= 'no '+reklam_ank+" "+anketa+" 1")
        keyboard_ask.add(button_accept)
        keyboard_ask.add(button_refusal)
        
        bot.send_message(get_id_po_ankete(reklam_ank,osn_base),"Один из покупателей(ID="+anketa+") подходит по параметрам и согласился с вами сотрудничать. Готовы ли вы? \nВот его данные:\n\n"+list_to_str_vivod(get_info_po_ankete(anketa,osn_base)),reply_markup=keyboard_ask)
        
    def test_in_main(num_id,address):
        wb = openpyxl.load_workbook(filename = address)
        sheet_ranges = wb['1']
        column_id = sheet_ranges['A']
        for i in range(1,len(column_id)):
            #print(column_id[i].value,num_id)        
            if str(column_id[i].value)==str(num_id):return True
        return False
    
    
            
    def obrabotka(anketa):
        wb1 = openpyxl.load_workbook(filename = all_base)
        wb2 = openpyxl.load_workbook(filename = osn_base)
        l=anketa.split(".")
        if l[0]=="0":
            sheet1 = wb1["0"]
            sheet2 = wb2["0"]
            num1=get_stroka_po_ankete(anketa,all_base)
            num2=(sheet2.cell(row=1, column=26)).value
            #print("num "+str(num2))
            sheet2['A'+str(num2)].value = sheet1['A'+str(num1)].value
            sheet2['B'+str(num2)].value = sheet1['B'+str(num1)].value
            sheet2['C'+str(num2)].value = sheet1['C'+str(num1)].value
            sheet2['D'+str(num2)].value = sheet1['D'+str(num1)].value
            sheet2['E'+str(num2)].value = sheet1['E'+str(num1)].value
            sheet2['F'+str(num2)].value = sheet1['F'+str(num1)].value
            sheet2['G'+str(num2)].value = sheet1['G'+str(num1)].value
            sheet2['H'+str(num2)].value = sheet1['H'+str(num1)].value
            sheet2['I'+str(num2)].value = sheet1['I'+str(num1)].value
            sheet2['J'+str(num2)].value = sheet1['J'+str(num1)].value
            sheet2['K'+str(num2)].value = sheet1['K'+str(num1)].value
            sheet2['Q'+str(num2)].value = sheet1['Q'+str(num1)].value
            sheet2["Z1"]=num2+1        
        else:
            sheet1 = wb1["1"]
            sheet2 = wb2["1"]
            num1=get_stroka_po_ankete(anketa,all_base)
            num2=(sheet2.cell(row=1, column=26)).value
            
            sheet2['A'+str(num2)].value = sheet1['A'+str(num1)].value
            sheet2['B'+str(num2)].value = sheet1['B'+str(num1)].value
            sheet2['C'+str(num2)].value = sheet1['C'+str(num1)].value
            sheet2['D'+str(num2)].value = sheet1['D'+str(num1)].value
            sheet2['E'+str(num2)].value = sheet1['E'+str(num1)].value
            sheet2['F'+str(num2)].value = sheet1['F'+str(num1)].value
            sheet2['G'+str(num2)].value = sheet1['G'+str(num1)].value
            sheet2['H'+str(num2)].value = sheet1['H'+str(num1)].value
            sheet2['Q'+str(num2)].value = sheet1['Q'+str(num1)].value
            sheet2["Z1"]=num2+1  
            
        wb1.save(all_base)
        wb2.save(osn_base) 
        
    def check_true_id(anketa,id_input,address):
        if str(get_id_po_ankete(anketa,address))==str(id_input):
            return True
        return False    
    
    def list_str__(l):
        s=""
        for i in range(len(l)-1):
            s=str(s+str(l[i])+"|")
        s=str(s+str(l[-1]))
        return s
    
    def from_list_to_str_1(l):
        s=""
        for i in range(len(l)):
            if i==1:
                s=str(s+"Социальная сеть: "+str(l[i])+"\n")
            if i==2:
                s=str(s+"Никнейм: "+str(l[i])+"\n")
            if i==3:
                s=str(s+"Подписчики: "+str(l[i])+"\n")
            if i==4:
                s=str(s+"Статистика: "+str(l[i])+"\n")
            if i==5:
                s=str(s+"Информация об аудитории: "+str(l[i])+"\n")
            #if i==6:
                #s=str(s+"Контактные данные: "+str(l[i])+"\n")    
            if i==9 and str(l[1])=="Instagram":
                s=str(s+str(inst(l[i]))+"\n")     
            elif i==9:
                s=str(s+"Цена: "+str(l[i])+"\n")
            if i==10:
                s=str(s+"Бартер: "+str(test_barter(l[i])))   
        return s
    def test_anketa(anketa):
        l=anketa.split(".")
        if len(l)!=2:
            return False
        if test_int(l[0])==False or test_int(l[1])==False:
            return False
        return True
    def contact(anketa):
        wb = openpyxl.load_workbook(filename = osn_base)
        sheet=wb['0']
        return sheet["G"+str(get_stroka_po_ankete(anketa,osn_base))].value
    
    def send_anketa(id_send,anketa,list_ank,number_ank_now):#not next!
        list_ank=list_ank.split("|")
        wb = openpyxl.load_workbook(filename = osn_base)
        sheet= wb['1']
        st=str(get_stroka_po_ankete(anketa,osn_base))
        
        if number_ank_now>=len(list_ank) or sheet["L"+st].value=="":
            bot.send_message(id_send,"Извините! На данный момент нет анкет для Вас. Повторите позднее или повторите запрос")
            sheet["L"+st].value=""
            sheet["M"+st].value=""   
            wb.save(osn_base)
        else:
            keyboard_ask = telebot.types.InlineKeyboardMarkup()
            button_accept = telebot.types.InlineKeyboardButton(text ="Согласен✅", callback_data = 'ok '+list_ank[number_ank_now]+" "+anketa)
            button_refusal = telebot.types.InlineKeyboardButton(text ="Не подходит❌", callback_data= 'no '+list_ank[number_ank_now]+" "+anketa)
            keyboard_ask.add(button_accept)
            keyboard_ask.add(button_refusal)
            bot.send_message(id_send,from_list_to_str_1(get_info_po_ankete(list_ank[number_ank_now],osn_base)),reply_markup=keyboard_ask)
    
    def podbor(message):
        if check_com(message)==0:
            inp_anketa=message.text
            if test_anketa(inp_anketa):
                if inp_anketa.split(".")[0]=="1":
                    if check_true_id(inp_anketa,message.from_user.id,osn_base):
                        if test_date(inp_anketa):
                            res=selection(inp_anketa,osn_base)
                            wb = openpyxl.load_workbook(filename = osn_base)
                            sheet= wb['1']        
                            #bot.send_message(get_id_po_ankete(inp_anketa,osn_base),"test!")
                            st=str(get_stroka_po_ankete(inp_anketa,osn_base))
                            if len(res)!=0:
                                bot.send_message(get_id_po_ankete(inp_anketa,osn_base),list_str__(res)) #test
                                q=list_str__(res)
                                sheet["L"+st].value=q
                                sheet["M"+st].value=0
                                wb.save(osn_base)
                                send_anketa(message.from_user.id,inp_anketa,q,0)
                            else:
                                bot.send_message(get_id_po_ankete(inp_anketa,osn_base),"Извините! На данный момент нет анкет для Вас. Повторите позднее")
                                sheet[str("L"+st)].value=None
                                sheet[str("M"+st)].value=None  
                                wb.save(osn_base)
                        else:
                            bot.send_message(message.from_user.id,"Срок действия вашей анкеты истек")
                    else:
                        bot.send_message(message.from_user.id,"Данная анкета не пренадлежит вам. Введите /search и попробуйте снова")
                else:
                    bot.send_message(message.from_user.id,"Рекламодателям нельзя использовать данную функцию. Введите /search и попробуйте снова")
            else:
                bot.send_message(message.from_user.id,"Вы ввели неправильную анкету. Введите /search и попробуйте снова")
        else:
            get_start_message(message)    
        
    def get_info_po_ankete(num_anketa,database):
        wb = openpyxl.load_workbook(filename = database)
        l=num_anketa.split(".")
        result=[]
        t=0
        if l[0]=="0":
            sheet = wb["0"]
            num = (sheet.cell(row=1, column=26)).value  
            for i in range(2,num):
                if sheet['Q'+str(i)].value==str(num_anketa):
                    t+=1
                    result.append(sheet['A'+str(i)].value)
                    result.append(sheet['B'+str(i)].value) 
                    result.append(sheet['C'+str(i)].value)
                    result.append(sheet['D'+str(i)].value)
                    result.append(sheet['E'+str(i)].value)
                    result.append(sheet['F'+str(i)].value) 
                    result.append(sheet['G'+str(i)].value) 
                    result.append(sheet['H'+str(i)].value)
                    result.append(sheet['I'+str(i)].value) 
                    result.append(sheet['J'+str(i)].value) 
                    result.append(sheet['K'+str(i)].value)   
                    result.append(sheet['Q'+str(i)].value) 
        if l[0]=="1":
            sheet = wb["1"]
            num = (sheet.cell(row=1, column=26)).value        
            for i in range(2,num):
                if sheet['Q'+str(i)].value==str(num_anketa): 
                    t+=1
                    result.append(sheet['A'+str(i)].value)
                    result.append(sheet['B'+str(i)].value) 
                    result.append(sheet['C'+str(i)].value)
                    result.append(sheet['D'+str(i)].value)
                    result.append(sheet['E'+str(i)].value)
                    result.append(sheet['F'+str(i)].value) 
                    result.append(sheet['G'+str(i)].value) 
                    result.append(sheet['H'+str(i)].value)   
                    result.append(sheet['Q'+str(i)].value) 
        if t!=0:
            return result
        else:
            return 0
        
    def selection(anketa_pokup,database):
        l_ank=anketa_pokup.split(".")
        
        
        lis=get_info_po_ankete(anketa_pokup,database)
        id_application_buyer_=lis[8]
        network_type_buyer_=lis[1]
        subscribers_buyer_=lis[3]
        statystics_buyer_=lis[4]
        name_buyer_=lis[2]
        info_buyer_=lis[5]
        category_ad_buyer_=lis[6]
        price_buyer_=lis[7]
        
        
        wb = openpyxl.load_workbook(filename = database)
        sheet_ranges = wb['0']
        column_id = sheet_ranges['A']
        column_category = sheet_ranges['I']
        column_price = sheet_ranges['J']
        column_barter = sheet_ranges['K']
        column_subs = sheet_ranges['D']
        column_name = sheet_ranges['C']
        column_info = sheet_ranges['F']
        column_stats = sheet_ranges['E']
        column_self_category =  sheet_ranges['H']
        
        result=[]
        
        sheet_ranges_1 = wb['0']
        num = (sheet_ranges_1.cell(row=1, column=26)).value
        for i in range(2,num-1):
            if str(sheet_ranges_1["N"+str(i)].value)=="1":
                if sheet_ranges_1["B"+str(i)].value==network_type_buyer_ and category_ad_buyer_ == column_category[i].value:
                    if check_price(price_buyer_,column_price[i].value):
                        result.append(sheet_ranges["Q"+str(i)].value)                
        return result    
        
    
    def inst(s):
        s1=""
        s=s.split(",")
        l=[]
        if s[0]!="-":
            l.append("Стоимость рекламного поста: "+s[0])
        if s[1]!="-":
            l.append("Стоимость сторис: "+s[1])
        if s[2]!="-":
            l.append("Стоимость вечного поста: "+s[2])
        for i in range(len(l)-1):
            s1=str(s1+str(l[i])+"\n")
        s1=str(s1+str(l[-1]))
        return s1    
        
    def test_barter(s):
        if s=="0":
            return "Бартер не интересует"
        else:
            return "Бартер интересует(Возможно с доплатой)"    
    """    
    def dob_anketa_contact(anketa):
        wb = openpyxl.load_workbook(filename = con_base)
        sheet=wb['main']
        stroka=0
        for i in range(2,int(sheet["Z1"].value)):
            if str(sheet["A"+str(i)].value)==str(get_id_po_ankete(anketa,all_base)):
                stroka=i
                
        if stroka!=0:
            s=str(sheet["D"+str(i)].value).split(",")
            if len(s)==1:
                s=[]
            s.append(anketa)
            q=""
            
            for i in range(len(s)-1):
                q=str(q)+str(s[i])+","
            q=str(q)+str(s[-1])
            sheet["D"+str(stroka)].value=q
            wb.save(con_base)    
    """
    def check_price(a,b):
        if test_int(b):
            if int(a)>=int(b):
                return True
            return False
        else:
            l=b.split(",")
        rr=0
        if l[0]!="-":
            if int(a)>=int(l[0]):
                rr+=1
        if l[1]!="-":
            if int(a)>=int(l[1]):
                rr+=1    
        if l[2]!="-":
            if int(a)>=int(l[2]):
                rr+=1    
        if rr!=0:
            return True
        return False
        
    def get_last_anketa_po_tgid(id):
        wb = openpyxl.load_workbook(filename = con_base)
        sheet= wb["main"] 
        a=int(sheet["Z1"].value)
        for i in range(2,a):
            if str(id)==str(sheet["A"+str(i)].value):
                return sheet["B"+str(i)].value
        return 0
    
    def test_int(s):
        try:
            s=int(s)
        except:
            return False
        return True
    
    def add_or_check_id(id,anketa):
        wb = openpyxl.load_workbook(filename = con_base)
        sheet= wb["main"]     
        if get_last_anketa_po_tgid(id)==0:
            #print("test!!!","B"+str(sheet["Z1"].value))
            sheet["A"+str(sheet["Z1"].value)].value=str(id)
            sheet["B"+str(sheet["Z1"].value)].value=str(anketa)
            sheet["Z1"].value=int(sheet["Z1"].value)+1
            wb.save(con_base)
        else:
            for i in range(2,int(sheet["Z1"].value)):
                if sheet["A"+str(i)].value==str(id):
                    sheet["B"+str(i)].value=anketa
                    wb.save(con_base)   
                    
    def get_stroka_po_ankete(anketa,database):
        l=anketa.split(".")
        wb = openpyxl.load_workbook(filename = database)
        if l[0]=="0":
            #print(0)
            sheet= wb["0"]     
            a=sheet["Z1"].value
            #print("a ", a)
            for i in range(2,int(a)):
                #print(sheet["Q"+str(i)].value,anketa)
                if str(sheet["Q"+str(i)].value)==str(anketa):
                    return str(i)
        if l[0]=="1":
            sheet= wb["1"]     
            a=sheet["Z1"].value
            for i in range(2,int(a)):
                if str(sheet["Q"+str(i)].value)==str(anketa):
                    return str(i)    
                
    def check_com(message):
        if message.text=="/start":
            return 1
        if message.text=="/search":
            return 2
        if message.text=="/info":
            return 3
        return 0
    def get_id_po_ankete(num_anketa,address):
        wb = openpyxl.load_workbook(filename = address)
        l=num_anketa.split(".")
        if l[0]=="0":
            sheet = wb["0"]
            num = (sheet.cell(row=1, column=26)).value        
            for i in range(2,int(num)):
                if sheet['Q'+str(i)].value==str(num_anketa): 
                    return (sheet.cell(row=i, column=1)).value 
        if l[0]=="1":
            sheet = wb["1"]
            num = (sheet.cell(row=1, column=26)).value        
            for i in range(2,int(num)):
                if sheet['Q'+str(i)].value==str(num_anketa):
                    return (sheet.cell(row=i, column=1)).value    
    def status(message):
        if check_com(message)==0:    
            anketa=message.text
            if test_anketa(anketa):
                if str(get_id_po_ankete(anketa,osn_base))==str(message.from_user.id):
                    wb = openpyxl.load_workbook(filename = osn_base)
                    sheet = wb[anketa.split(".")[0]]
                
                    if test_date(anketa)==False:
                        bot.send_message(message.from_user.id,"Ваша анкета действительна до "+sheet["U"+str(get_stroka_po_ankete(anketa,osn_base))].value+"\nСрок действия анкеты Вашей анкеты закончился. Для продления обратитесь к администратору(@metand_2)\nДля повтора нажмите /start")
                    else:
                        bot.send_message(message.from_user.id,"Ваша анкета действительна до "+sheet["U"+str(get_stroka_po_ankete(anketa,osn_base))].value+"\nДля повтора нажмите /start")
                else:
                    bot.send_message(message.from_user.id,"Данная анкета не пренадлежит вам\nДля повтора нажмите /start")
            else:
                bot.send_message(message.from_user.id,"Вы ввели неверный формат анкеты")
        else:
            get_start_message(message)
    #-----------------------------------------------------------------------------------------------------------Время
    #print(time.ctime(time.time()+56000))     Mon Jun  1 17:05:53 2020
    def sec_to_date(time_,delta=0):
        #a=time.time()
        ts = int(time_+delta)
        return datetime.datetime.fromtimestamp(ts).strftime("%d.%m.%Y")
    
    def test_date(anketa): #до какого времени
        l=anketa.split(".")
        wb = openpyxl.load_workbook(osn_base)
        sheet=wb[l[0]]
        #print(anketa)
        #print(get_stroka_po_ankete(anketa,osn_base))
        a=sheet["T"+get_stroka_po_ankete(anketa,osn_base)].value
        a=float(a)
        ts=time.time()
        if ts<=a:
            return True
        return False
    
    #-----------------------------------------------------------------------------------------------------------Рабочие функции
    def plus(anketa,time_):
        wb = openpyxl.load_workbook(filename = osn_base)
        sheet = wb[anketa.split(".")[0]]    
    
        sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value=float(sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value)+86400*float(time_)*30
        sheet["U"+str(get_stroka_po_ankete(anketa,osn_base))].value=sec_to_date(sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value)
        wb.save(osn_base)
    def minus(anketa,time_):
        wb = openpyxl.load_workbook(filename = osn_base)
        sheet = wb[anketa.split(".")[0]]    
    
        sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value=float(sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value)-86400*int(time_)*30
        sheet["U"+str(get_stroka_po_ankete(anketa,osn_base))].value=sec_to_date(sheet["T"+str(get_stroka_po_ankete(anketa,osn_base))].value)
        wb.save(osn_base)    
    def plus_check(s):
        if len(s.split("+"))==2:
            return True
        return False
    def minus_check(s):
        if len(s.split("-"))==2:
            return True
        return False
    #определение типа клиента    
    def get_type_message(message):
        if check_com(message)==0:
            if message.text == 'Рекламодатель':
                wb = openpyxl.load_workbook(filename= all_base)
                sheet=wb["0"]                 
                a=int(sheet["Z1"].value)
                anketa="0."+str(a)
    
                sheet["A"+str(a)].value=str(message.from_user.id)
                sheet["Q"+str(a)].value=anketa
                sheet["Z1"].value=str(a+1)
                wb.save(all_base)
                
                add_or_check_id(message.from_user.id,anketa)
                
                wb1 = openpyxl.load_workbook(filename= con_base)
                sheet1=wb1["main"]       
                #print(sheet1["B2"].value)
                bot.send_message(message.from_user.id, "Пожалуйста, напишите, в какой социальной сети Вы готовы разместить рекламу. ", reply_markup=keyboard_network_type)
                bot.register_next_step_handler(message, get_network_type_advertiser)
            elif message.text == 'Покупатель':
                
                wb = openpyxl.load_workbook(filename = all_base)
                sheet=wb["1"]                 
                a=int(sheet["Z1"].value)
                anketa="1."+str(a)
                sheet["A"+str(a)].value=str(message.from_user.id)
                sheet["Q"+str(a)].value=anketa
                sheet["Z1"].value=str(a+1)
                wb.save(all_base)
                add_or_check_id(message.from_user.id,anketa)
                
                bot.send_message(message.from_user.id, "Пожалуйста, напишите, в какой социальной сети Вы хотите купить рекламу. ", reply_markup=keyboard_network_type)
                bot.register_next_step_handler(message, get_network_type_buyer)
        else:
            get_start_message(message)
            
            
     #ниже получаем название соц сети
    def get_network_type_advertiser(message): 
        if check_com(message)==0:
            network_type_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["B"+get_stroka_po_ankete(anketa,all_base)].value=network_type_advertiser
            wb.save(all_base)
            
            bot.send_message(message.from_user.id, "Введите никнейм вашего канала / аккаунта / профиля.", reply_markup = types.ReplyKeyboardRemove(selective=False))
            bot.register_next_step_handler(message, get_name_advertiser)
        else:
            get_start_message(message)     
    def get_network_type_buyer(message):
        if check_com(message)==0:
            network_type_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["B"+get_stroka_po_ankete(anketa,all_base)].value=network_type_buyer
            wb.save(all_base)        
            
            bot.send_message(message.from_user.id, "Введите никнейм вашего канала / аккаунта / профиля.", reply_markup = types.ReplyKeyboardRemove(selective=False))
            bot.register_next_step_handler(message, get_name_buyer)
        else:
            get_start_message(message)         
            
            
    #ниже получаем никнейм
    def get_name_advertiser(message):
        if check_com(message)==0:
            name_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["C"+get_stroka_po_ankete(anketa,all_base)].value=name_advertiser
            wb.save(all_base)            
            
            bot.send_message(message.from_user.id, "Введите количество подпиcчиков у вашего канала / аккаунта / профиля.")
            bot.register_next_step_handler(message, get_subscribers_advertiser)
        else:
            get_start_message(message)     
    def get_name_buyer(message):
        if check_com(message)==0:
            name_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["C"+get_stroka_po_ankete(anketa,all_base)].value=name_buyer
            wb.save(all_base)         
            
            bot.send_message(message.from_user.id, "Введите количество подпиcчиков у вашего канала / аккаунта / профиля.")
            bot.register_next_step_handler(message, get_subscribers_buyer)
        else:
            get_start_message(message)        
            
    #ниже получаем количество подписчиков
    def get_subscribers_advertiser(message):
        if check_com(message)==0:
            subscribers_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["D"+get_stroka_po_ankete(anketa,all_base)].value=subscribers_advertiser
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите среднюю посещаемость/просмотров/лайков за неделю.")
            bot.register_next_step_handler(message, get_statystics_advertiser)
        else:
            get_start_message(message)     
            
    def get_subscribers_buyer(message):
        if check_com(message)==0:
            subscribers_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["D"+get_stroka_po_ankete(anketa,all_base)].value=subscribers_buyer
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите среднюю посещаемость/просмотров/лайков за неделю.")
            bot.register_next_step_handler(message, get_statystics_buyer)
        else:
            get_start_message(message)  
            
    #ниже получаем среднюю статистику
    def get_statystics_advertiser(message):
        if check_com(message)==0:
            statystics_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["E"+get_stroka_po_ankete(anketa,all_base)].value=statystics_advertiser
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите страну, город, пол, возраст вашей аудитории.")
            bot.register_next_step_handler(message, get_info_advertiser)  
        else:
            get_start_message(message)      
    def get_statystics_buyer(message):
        if check_com(message)==0:
            statystics_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["E"+get_stroka_po_ankete(anketa,all_base)].value=statystics_buyer
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите страну, город, пол, возраст вашей аудитории.")
            bot.register_next_step_handler(message, get_info_buyer)
        else:
            get_start_message(message)          
        
         
    #ниже получаем данные об аудитории
    def get_info_advertiser(message):
        if check_com(message)==0:
            info_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["F"+get_stroka_po_ankete(anketa,all_base)].value=info_advertiser
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите Ваши контактные данные для связи насчет рекламы(Telegram, Instagram, email и т.д.)")
            bot.register_next_step_handler(message, get_contacts_advertiser)
        else:
            get_start_message(message)              
    def get_info_buyer(message): 
        if check_com(message)==0:
            info_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["F"+get_stroka_po_ankete(anketa,all_base)].value=info_buyer
            wb.save(all_base)          
            
            bot.send_message(message.from_user.id, "Укажите категорию вашей рекламы.", reply_markup = keyboard_category_ad)
            bot.register_next_step_handler(message, get_category_buyer)
        else:
            get_start_message(message)          
        
    #ниже получаем контакты рекламодателя
    def get_contacts_advertiser(message):
        if check_com(message)==0:
            contacts_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["G"+get_stroka_po_ankete(anketa,all_base)].value=contacts_advertiser
            wb.save(all_base)                  
            
            bot.send_message(message.from_user.id, "Укажите категорию вашего канала / аккаунта / профиля.", reply_markup=keyboard_category_ad)
            bot.register_next_step_handler(message, get_self_category_advertiser)
        else:
            get_start_message(message)  
    
    #ниже получаем категорию канала рекламодателя    
    def get_self_category_advertiser(message):
        if check_com(message)==0:
            self_category_ad = str(message.text)
    
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["H"+get_stroka_po_ankete(anketa,all_base)].value=self_category_ad
            wb.save(all_base)                  
            
            bot.send_message(message.from_user.id, "📄 Категория рекламы: ", reply_markup=keyboard_category_ad)
            bot.register_next_step_handler(message, get_category_advertiser)
        else:
            get_start_message(message)        
    
    #реклама инстаграма
    def price_adv_1(message):
        if check_com(message)==0:
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]            
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["J"+get_stroka_po_ankete(anketa,all_base)].value=message.text
            wb.save(all_base)
            bot.send_message(message.from_user.id, "💰 Введите стоимость рекламной сторис(пожалуйста, вводите цену в рублях\nEсли же вы не предоставляете рекламы данного типа, то вместо стоимости введите 0)",reply_markup = types.ReplyKeyboardRemove(selective=False))
            bot.register_next_step_handler(message, price_adv_2)
        else:
            get_start_message(message)
    def price_adv_2(message):
        if check_com(message)==0:
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]            
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["J"+get_stroka_po_ankete(anketa,all_base)].value=str(str(sheet["J"+get_stroka_po_ankete(anketa,all_base)].value)+","+str(message.text))
            wb.save(all_base)
            bot.send_message(message.from_user.id, "💰 Введите стоимость постоянного рекламного поста(пожалуйста, вводите цену в рублях\nEсли же вы не предоставляете рекламы данного типа, то вместо стоимости введите 0)",reply_markup = types.ReplyKeyboardRemove(selective=False))
            bot.register_next_step_handler(message, price_adv_3)
        else:
            get_start_message(message)
    def price_adv_3(message):
        if check_com(message)==0:
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]            
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["J"+get_stroka_po_ankete(anketa,all_base)].value=str(str(sheet["J"+get_stroka_po_ankete(anketa,all_base)].value)+","+str(message.text))
            wb.save(all_base)
            l=sheet["J"+get_stroka_po_ankete(anketa,all_base)].value.split(",")
            s=""
            for i in range(len(l)):
                if l[i]=="0":
                    l[i]="-"
            for i in range(len(l)-1):
                s=s+str(l[i])+","
            s=s+str(l[-1])
            sheet["J"+get_stroka_po_ankete(anketa,all_base)].value=s
            wb.save(all_base)        
            print(sheet["J"+get_stroka_po_ankete(anketa,all_base)].value)
            
            bot.send_message(message.from_user.id, "♻️ Если вы готовы на бартер/коллаборацию, введите максимальное отличие подписчиков между Вами и покупателем, если нет, то введите 0")
            bot.register_next_step_handler(message, application_advertiser)
            """
            bot.send_message(message.from_user.id, "♻️ Если вы готовы на бартер/коллаборацию, введите максимальное отличие подписчиков между Вами и покупателем, если нет, то введите 0")
            bot.register_next_step_handler(message, application_advertiser)
            """
        else:
            get_start_message(message)
    #ниже получим категорию рекламы
    def get_category_advertiser(message):
        if check_com(message)==0:
            category_ad_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]            
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["I"+get_stroka_po_ankete(anketa,all_base)].value=category_ad_advertiser
                   
            network_type_advertiser=sheet["B"+get_stroka_po_ankete(anketa,all_base)].value
            type_=0
            if network_type_advertiser == "Instagram":
                type_ = 3
                #bot.send_message(message.from_user.id, "💰 Введите стоимость рекламного поста, рекламной сторис, постоянного рекламного поста(пожалуйста, вводите цены в рублях и через запятую; если вы не предоставляете рекламы данного типа, то вместо стоимости введите 0. Например, '0,0,0' или '0,1000,1590')",reply_markup = types.ReplyKeyboardRemove(selective=False))
                bot.send_message(message.from_user.id, "💰 Введите стоимость временного рекламного поста(пожалуйста, вводите цену в рублях\nEсли же вы не предоставляете рекламы данного типа, то вместо стоимости введите 0)",reply_markup = types.ReplyKeyboardRemove(selective=False))
                sheet["L"+get_stroka_po_ankete(anketa,all_base)].value=type_
                wb.save(all_base)   
                bot.register_next_step_handler(message, price_adv_1)            
            elif network_type_advertiser == "Telegram":
                type_ = 1
                bot.send_message(message.from_user.id, "💰 Введите стоимость рекламного поста (пожалуйста, вводите цены в рублях)",reply_markup = types.ReplyKeyboardRemove(selective=False))
                sheet["L"+get_stroka_po_ankete(anketa,all_base)].value=type_
                
                wb.save(all_base)   
                bot.register_next_step_handler(message, price_adv)            
            elif network_type_advertiser == "YouTube":
                type_ = 1
                bot.send_message(message.from_user.id, "💰 Введите стоимость упоминания рекламного характера в видео(пожалуйста, вводите цены в рублях)",reply_markup = types.ReplyKeyboardRemove(selective=False))
                sheet["L"+get_stroka_po_ankete(anketa,all_base)].value=type_
                
                wb.save(all_base)   
                bot.register_next_step_handler(message, price_adv)            
            elif network_type_advertiser == "Tik-Tok":
                type_ = 1
                bot.send_message(message.from_user.id, "💰 Введите стоимость рекламной интеграциии(пожалуйста, вводите цену в рублях)",reply_markup = types.ReplyKeyboardRemove(selective=False))
                sheet["L"+get_stroka_po_ankete(anketa,all_base)].value=type_
                
                wb.save(all_base)   
                bot.register_next_step_handler(message, price_adv)            
            elif network_type_advertiser == "Вконтакте":
                type_ = 1
                bot.send_message(message.from_user.id, "💰 Введите стоимость рекламного поста(пожалуйста, вводите цену в рублях)",reply_markup = types.ReplyKeyboardRemove(selective=False))
                sheet["L"+get_stroka_po_ankete(anketa,all_base)].value=type_
                
                wb.save(all_base)   
                bot.register_next_step_handler(message, price_adv)            
            if type_==0:
                bot.send_message(message.from_user.id, "Вы ввели ошибочные данные. Начните сначала. Введите /start")
                return 0
        else:
            get_start_message(message)           
            
    def get_category_buyer(message):
        if check_com(message)==0:
            
            category_ad_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["G"+get_stroka_po_ankete(anketa,all_base)].value=category_ad_buyer
            wb.save(all_base)                 
            
            bot.send_message(message.from_user.id, "💰 Укажите максимальную сумму, за которую вы желаете приобрести рекламу")
            bot.register_next_step_handler(message, application_buyer)
        else:
            get_start_message(message)             
    
    
    
    # получим стоимость рекламной кампании и возможность бартера
    def price_adv(message):
        if check_com(message)==0:
            price_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]   
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            type_=sheet["L"+get_stroka_po_ankete(anketa,all_base)].value        
            
            if type_ == 3:
                price_advertiser = price_advertiser.split(",")
                price_advertiser=list(price_advertiser)
                for i in range(len(price_advertiser)):
                    #price_advertiser[i]=int(price_advertiser[i])
                    if int(price_advertiser[i])==0:
                        price_advertiser[i] = '-'    
                price_advertiser=str(str(price_advertiser[0])+","+str(price_advertiser[1])+","+str(price_advertiser[2]))
            sheet["J"+get_stroka_po_ankete(anketa,all_base)].value=price_advertiser
            wb.save(all_base)
            bot.send_message(message.from_user.id, "♻️ Если вы готовы на бартер/коллаборацию, введите максимальное отличие подписчиков между Вами и покупателем, если нет, то введите 0")
            bot.register_next_step_handler(message, application_advertiser)
        else:
            get_start_message(message) 
            
            
    def application_buyer(message):
        if check_com(message)==0:
            price_buyer = str(message.text)   
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["H"+get_stroka_po_ankete(anketa,all_base)].value=price_buyer
            wb.save(all_base)                    
            
                 
            
            stroka=get_stroka_po_ankete(anketa,all_base)
            network_type_buyer = sheet["B"+stroka].value
            name_buyer = sheet["C"+stroka].value
            subscribers_buyer = sheet["D"+stroka].value
            statystics_buyer = sheet["E"+stroka].value
            info_buyer = sheet["F"+stroka].value
            category_ad_buyer = sheet["G"+stroka].value
            
            #s="📝 Ваша анкета: " + '\n' + "🌐 1.Социальная сеть: " + network_type_buyer + '\n' + "👨‍💻 2.Никнейм: " + name_buyer + '\n'  "👥 3."+ subscribers_buyer + "подписчиков"+ '\n' + "📊 4.Статистика профиля: " + statystics_buyer + '\n' + "🗺 5.Инфомация об аудитории: " + info_buyer + '\n' + "📄 6.Категория рекламы: " + category_ad_buyer + '\n' + "💰 7. Максимальная стоимость рекламы: "+ price_buyer + '\n' +'\n' + "Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." + '\n'  + "Ваша анкета составлена верно?"
            
            keyboard_answer = telebot.types.InlineKeyboardMarkup() #кнопки для подтверждения правильности анкеты клиентом
            button_yes = telebot.types.InlineKeyboardButton(text ="Да, все верно✅", callback_data = 'yes '+str(anketa))    
            button_no = telebot.types.InlineKeyboardButton(text ="Нет, изменить❌", callback_data = 'no '+str(anketa))   
            keyboard_answer.add(button_yes)#добавление кнопок в набор
            keyboard_answer.add(button_no)  
            
            bot.send_message(message.from_user.id, 
                             "📝 Ваша анкета: " + '\n' 
                             + "🌐 1.Социальная сеть: " + network_type_buyer + '\n' 
                             + "👨‍💻 2.Никнейм: " + name_buyer + '\n' 
                             + "👥 3."+ subscribers_buyer + "подписчиков"+ '\n' 
                             + "📊 4.Статистика профиля: " + statystics_buyer + '\n' 
                             + "🗺 5.Инфомация об аудитории: " + info_buyer + '\n' 
                             + "📄 6.Категория рекламы: " + category_ad_buyer + '\n' 
                             + "💰 7. Максимальная стоимость рекламы: "+ price_buyer + '\n' +'\n' 
                             + "Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." + '\n' 
                             + "Ваша анкета составлена верно?", 
                             reply_markup=keyboard_answer)
        else:
            get_start_message(message) 
            
            #   #######
            
            
    def application_advertiser(message):
        if check_com(message)==0:
            barter_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["K"+get_stroka_po_ankete(anketa,all_base)].value=barter_advertiser
            wb.save(all_base)             
            
            keyboard_answer = telebot.types.InlineKeyboardMarkup() #кнопки для подтверждения правильности анкеты клиентом
            button_yes = telebot.types.InlineKeyboardButton(text ="Да, все верно✅", callback_data = 'yes '+str(anketa))    
            button_no = telebot.types.InlineKeyboardButton(text ="Нет, изменить❌", callback_data = 'no '+str(anketa))   
            keyboard_answer.add(button_yes)# добавление кнопок в набор
            keyboard_answer.add(button_no)        
            
            stroka=get_stroka_po_ankete(anketa,all_base)
            
            type_=int(sheet["L"+stroka].value)
            network_type_advertiser=sheet["B"+stroka].value
            name_advertiser=sheet["C"+stroka].value
            subscribers_advertiser=sheet["D"+stroka].value
            statystics_advertiser=sheet["E"+stroka].value
            info_advertiser=sheet["F"+stroka].value
            contacts_advertiser=sheet["G"+stroka].value
            self_category_ad=sheet["H"+stroka].value
            category_ad_advertiser=sheet["I"+stroka].value
            price_advertiser=sheet["J"+stroka].value
            
            #print(price_advertiser)
            if type_==3:
                qwe=price_advertiser.split(",")
            if int(barter_advertiser)!=0:
                if type_ == 3:
                    bot.send_message(message.from_user.id, 
                                     "📝 Ваша анкета: " + '\n' 
                                     + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' 
                                     + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' 
                                     + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' 
                                     + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' 
                                     + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'
                                     + "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' 
                                     + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ qwe[0] + '\n' + "💰 Стоимость сторис: " +qwe[1]+'\n' + "💰 Стоимость вечного поста: " + qwe[2] + '\n' 
                                     + "♻️ 10. Возможность бартера при разнице подписчиков: "+str(barter_advertiser)+'\n'+'\n' 
                                     +"Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." + '\n' +"Будущим клиентам будет видна ваша анкета кроме контактов. " + '\n'+ "Ваша анкета составлена верно?", reply_markup=keyboard_answer)  
                elif type_ == 1:
                    bot.send_message(message.from_user.id, 
                                     "📝 Ваша анкета: " + '\n' 
                                     + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' 
                                     + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' 
                                     + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' 
                                     + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' 
                                     + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'
                                     + "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' 
                                     + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ price_advertiser + '\n' 
                                     + "♻️ 10. Возможность бартера при разнице подписчиков: "+  barter_advertiser+'\n'+'\n' 
                                     + "Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." + '\n' 
                                     +"Будущим клиентам будет видна ваша анкета кроме контактов. " + '\n'
                                     + "Ваша анкета составлена верно?", 
                                     reply_markup=keyboard_answer)  
            else:
                if type_ == 3:
                    bot.send_message(message.from_user.id, 
                                     "📝 Ваша анкета: " + '\n' 
                                     + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' 
                                     + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' 
                                     + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' 
                                     + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' 
                                     + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'
                                     + "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' 
                                     + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ qwe[0] + '\n' + "💰 Стоимость сторис: " +qwe[1]+'\n' + "💰 Стоимость вечного поста: " + qwe[2] + '\n' 
                                     +"Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." + '\n' 
                                     +"Будущим клиентам будет видна ваша анкета кроме контактов. " + '\n'
                                     + "Ваша анкета составлена верно?", 
                                     reply_markup=keyboard_answer)  
                elif type_ == 1:
                    bot.send_message(message.from_user.id, 
                                     "📝 Ваша анкета: " + '\n' 
                                     + "🌐 1.Социальная сеть: " + network_type_advertiser + '\n' 
                                     + "👨‍💻 2.Никнейм: " + name_advertiser + '\n' 
                                     + "👥 3.Количество подписчиков: " + subscribers_advertiser + '\n' 
                                     + "📊 4.Статистика профиля: " + statystics_advertiser + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + info_advertiser + '\n' 
                                     + "📱 6.Ваши контакты: " + contacts_advertiser + '\n'
                                     + "📎  7.Категория Вашего профиля: " + self_category_ad + '\n' 
                                     + "📄 8.Категория рекламы: " + category_ad_advertiser + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ price_advertiser + '\n' + '\n' 
                                     + "Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." 
                                     + '\n' +"Будущим клиентам будет видна ваша анкета кроме контактов. " + '\n'
                                     + "Ваша анкета составлена верно?", 
                                     reply_markup=keyboard_answer)
        else:
            get_start_message(message)                     
    # --------------------------------------------------
    #функция для отправки анкеты админам, выше функции, которая вызывает данную
    def sending_application(id_anketa):
        network_type=int(id_anketa.split(".")[0])
        
        #print("i send application")
        keyboard_admin = telebot.types.InlineKeyboardMarkup()
        
        button_accept = telebot.types.InlineKeyboardButton(text ="Одобрить анкету✅", callback_data = 'accept '+id_anketa)
        button_refusal = telebot.types.InlineKeyboardButton(text ="Отклонить анкету❌", callback_data= 'refusal '+id_anketa)
        
        keyboard_admin.add(button_accept)
        keyboard_admin.add(button_refusal)
        
        if network_type == 0: #вывод анкеты рекламодателя
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(get_id_po_ankete(id_anketa,all_base))
             
            
            stroka=get_stroka_po_ankete(id_anketa,all_base)
            
            type_=int(sheet["L"+stroka].value)
            network_type_advertiser=sheet["B"+stroka].value
            name_advertiser=sheet["C"+stroka].value
            subscribers_advertiser=sheet["D"+stroka].value
            statystics_advertiser=sheet["E"+stroka].value
            info_advertiser=sheet["F"+stroka].value
            contacts_advertiser=sheet["G"+stroka].value
            self_category_ad=sheet["H"+stroka].value
            category_ad_advertiser=sheet["I"+stroka].value
            price_advertiser=sheet["J"+stroka].value
            barter_advertiser=sheet["K"+stroka].value
            
            bot_checker.send_message(405027580, 
                                     
                                     "Заявка №"+str(id_anketa)+'\n'
                                     +"Тип клиента: Рекламодатель" + '\n' 
                                     + "🌐 1.Социальная сеть:" + str(network_type_advertiser) + '\n' 
                                     + "👨‍💻 2.Никнейм: : " + str(name_advertiser) + '\n' 
                                     + "👥 3.Количество подписчиков: " + str(subscribers_advertiser) + '\n' 
                                     + "📊 4.Статистика профиля: " + str(statystics_advertiser) + '\n' 
                                     + "🗺 5.Инфомация об аудитории:  " + str(info_advertiser) + '\n' 
                                     + "📱 6.Ваши контакты: " + str(contacts_advertiser) + '\n' 
                                     + "📎 7.Категория Вашего профиля: " + str(self_category_ad) + '\n' 
                                     + "📄 8.Категория рекламы: " + str(category_ad_advertiser) + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ str(price_advertiser) + '\n'
                                     +"Возможность бартера при разнице подписчиков: "+  str(barter_advertiser)+'\n',
                                     reply_markup=keyboard_admin)
            
            bot_checker.send_message(741710024, 
                                     "Заявка №"+str(id_anketa)+'\n'
                                     +"Тип клиента: Рекламодатель" + '\n' 
                                     + "🌐 1.Социальная сеть:" + str(network_type_advertiser) + '\n' 
                                     + "👨‍💻 2.Никнейм: : " + str(name_advertiser) + '\n' 
                                     + "👥 3.Количество подписчиков: " + str(subscribers_advertiser) + '\n' 
                                     + "📊 4.Статистика профиля: " + str(statystics_advertiser) + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + str(info_advertiser) + '\n' 
                                     + "📱 6.Ваши контакты: " + str(contacts_advertiser) + '\n' 
                                     + "📎 7.Категория Вашего профиля: " + str(self_category_ad) + '\n' 
                                     + "📄 8.Категория рекламы: " + str(category_ad_advertiser) + '\n' 
                                     + "💰 9. Стоимость рекламного поста: "+ str(price_advertiser) + '\n'
                                     +"Возможность бартера при разнице подписчиков: "+  str(barter_advertiser)+'\n',
                                     reply_markup=keyboard_admin )
            
        else: 
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(get_id_po_ankete(id_anketa,all_base))
            
            stroka=get_stroka_po_ankete(anketa,all_base)
            
            network_type_buyer = sheet["B"+stroka].value
            name_buyer = sheet["C"+stroka].value
            subscribers_buyer = sheet["D"+stroka].value
            statystics_buyer = sheet["E"+stroka].value
            info_buyer = sheet["F"+stroka].value
            category_ad_buyer = sheet["G"+stroka].value  
            price_buyer=sheet["H"+stroka].value
            
            bot_checker.send_message(741710024, 
                                     "Заявка №"+str(id_anketa)+'\n'
                                     +"Тип клиента: покупатель рекламы" + '\n' 
                                     + "🌐 1.Социальная сеть:" + str(network_type_buyer) + '\n' 
                                     + "👨‍💻 2.Никнейм: : " + str(name_buyer) + '\n' 
                                     + "👥 3.Количество подписчиков: " + str(subscribers_buyer) + '\n' 
                                     + "📊 4.Статистика профиля: " + str(statystics_buyer) + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + str(info_buyer) + '\n' 
                                     + "📄 6.Категория рекламы: " + str(category_ad_buyer) + '\n'
                                     + "💰 7. Максимальная стоимость рекламы: "+ str(price_buyer) + '\n',
                                     reply_markup=keyboard_admin)
            
            bot_checker.send_message(405027580, 
                                     "Заявка №"+str(id_anketa)+'\n'
                                     +"Тип клиента: покупатель рекламы" + '\n' 
                                     + "🌐 1.Социальная сеть:" + str(network_type_buyer) + '\n' 
                                     + "👨‍💻 2.Никнейм: : " + str(name_buyer) + '\n' 
                                     + "👥 3.Количество подписчиков: " + str(subscribers_buyer) + '\n' 
                                     + "📊 4.Статистика профиля: " + str(statystics_buyer) + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + str(info_buyer) + '\n' 
                                     + "📄 6.Категория рекламы: " + str(category_ad_buyer) + '\n'
                                     + "💰 7. Максимальная стоимость рекламы: "+ str(price_buyer) + '\n',
                                     reply_markup=keyboard_admin)
            
    #ниже функции по изменению анкеты клиентом       
    def number_edit_advertiser(message):
        if check_com(message)==0:
        
            number_advertiser = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["M"+get_stroka_po_ankete(anketa,all_base)].value=number_advertiser
            wb.save(all_base)              
            
            bot.send_message(message.from_user.id, "Внесите изменения.")
            bot.register_next_step_handler(message, editing_advertiser)
        else:
            get_start_message(message)    
    def number_edit_buyer(message):
        if check_com(message)==0:
            number_buyer = str(message.text)
            
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            sheet["M"+get_stroka_po_ankete_all(anketa)].value=number_advertiser
            wb.save(all_base)           
            
            bot.send_message(message.from_user.id, "Внесите изменения.")
            bot.register_next_step_handler(message, editing_buyer)
        else:
            get_start_message(message)
    def editing_advertiser(message):
        #print("test edit_adv")
        if check_com(message)==0:
    
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["0"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            wb.save(all_base)                    
            
            keyboard_answer = telebot.types.InlineKeyboardMarkup() #кнопки для подтверждения правильности анкеты клиентом
            button_yes = telebot.types.InlineKeyboardButton(text ="Да, все верно✅", callback_data = 'yes '+str(anketa))    
            button_no = telebot.types.InlineKeyboardButton(text ="Нет, изменить❌", callback_data = 'no '+str(anketa))   
            keyboard_answer.add(button_yes)# добавление кнопок в набор
            keyboard_answer.add(button_no)   
            
            stroka=get_stroka_po_ankete_all(anketa) 
            
            network_type_advertiser=sheet["B"+stroka].value
            name_advertiser=sheet["C"+stroka].value
            subscribers_advertiser=sheet["D"+stroka].value
            statystics_advertiser=sheet["E"+stroka].value
            info_advertiser=sheet["F"+stroka].value
            contacts_advertiser=sheet["G"+stroka].value
            self_category_ad=sheet["H"+stroka].value
            category_ad_advertiser=sheet["I"+stroka].value
            price_advertiser=sheet["J"+stroka].value        
            barter_advertiser=sheet["K"+stroka].value
             
            number_advertiser=sheet["M"+stroka].value
            
            edit_advertiser = message.text
            if test_int(number_advertiser):
                if int(number_advertiser)>=1 and int(number_advertiser)<=8:
                    if number_advertiser == '1':
                        sheet["B"+stroka].value = edit_advertiser
                    if number_advertiser == '2':
                        sheet["C"+stroka].value=edit_advertiser
                    if number_advertiser == '3':
                        sheet["D"+stroka].value=edit_advertiser
                    if number_advertiser == '4':
                        sheet["E"+stroka].value = edit_advertiser
                    if number_advertiser == '5':
                        sheet["F"+stroka].value=edit_advertiser
                    if number_advertiser == '6':
                        sheet["G"+stroka].value=edit_advertiser
                    if number_advertiser == '7':
                        sheet["H"+stroka].value=edit_advertiser
                    if number_advertiser == '8':
                        sheet["I"+stroka].value=edit_advertiser
                    if number_advertiser == '9':
                        sheet["J"+stroka].value=edit_advertiser
                    if number_advertiser == '10':
                        sheet["K"+stroka].value=edit_advertiser
                    wb.save(all_base)
                    
                    if int(barter_advertiser)==0:
                        bot.send_message(message.from_user.id, 
                                         "Ваша анкета: " + '\n' 
                                         + "🌐 1.Социальная сеть:" + sheet["B"+stroka].value + '\n' 
                                         + "👨‍💻 2.Никнейм: " + sheet["C"+stroka].value + '\n' 
                                         + "👥 3.Количество подписчиков: " + sheet["D"+stroka].value + '\n' 
                                         + "📊 4.Статистика профиля: " + sheet["E"+stroka].value + '\n' 
                                         + "🗺 5.Инфомация об аудитории: " + sheet["F"+stroka].value + '\n' 
                                         + "📱 6.Ваши контакты: " + sheet["G"+stroka].value + '\n' 
                                         + "📎 7.Категория Вашего профиля: " + sheet["H"+stroka].value + '\n' 
                                         + "📄 8.Категория рекламы: " + sheet["I"+stroka].value + '\n'
                                         + "9.Цена рекламы:"+sheet["J"+stroka].value+'\n'
                                         + "Ваша анкета составлена верно?", 
                                         reply_markup=keyboard_answer)
                        
                    else:
                        bot.send_message(message.from_user.id, 
                                         "Ваша анкета: " + '\n' 
                                         + "🌐 1.Социальная сеть:" + sheet["B"+stroka].value + '\n' 
                                         + "👨‍💻 2.Никнейм: " + sheet["C"+stroka].value + '\n' 
                                         + "👥 3.Количество подписчиков: " + sheet["D"+stroka].value + '\n' 
                                         + "📊 4.Статистика профиля: " + sheet["E"+stroka].value + '\n' 
                                         + "🗺 5.Инфомация об аудитории: " + sheet["F"+stroka].value + '\n' 
                                         + "📱 6.Ваши контакты: " + sheet["G"+stroka].value + '\n' 
                                         + "📎 7.Категория Вашего профиля: " + sheet["H"+stroka].value + '\n' 
                                         + "📄 8.Категория рекламы: " + sheet["I"+stroka].value + '\n'
                                         + "9.Цена рекламы:"+sheet["J"+stroka].value+'\n'
                                         + "10.Максимальное отличие подписчиков"+sheet["K"+stroka].value+'\n'
                                         + "Ваша анкета составлена верно?", 
                                         reply_markup=keyboard_answer)
                    #bot.register_next_step_handler(message, checking_advertiser)   
                else:
                    bot.send_message(message.from_user.id,"Ошибка! Введите пункт еще раз!")
                    bot.register_next_step_handler(message, number_edit_advertiser)  
            else:
                bot.send_message(message.from_user.id,"Ошибка! Введите пункт еще раз!")
                bot.register_next_step_handler(message, number_edit_advertiser)      
            
        else:
            get_start_message(message)            
            
    def editing_buyer(message):
        if check_com(message)==0:
        
            wb = openpyxl.load_workbook(filename = all_base)
            sheet=wb["1"]    
            anketa=get_last_anketa_po_tgid(message.from_user.id)
            
            keyboard_answer = telebot.types.InlineKeyboardMarkup() #кнопки для подтверждения правильности анкеты клиентом
            button_yes = telebot.types.InlineKeyboardButton(text ="Да, все верно✅", callback_data = 'yes '+str(anketa))    
            button_no = telebot.types.InlineKeyboardButton(text ="Нет, изменить❌", callback_data = 'no '+str(anketa))   
            keyboard_answer.add(button_yes)# добавление кнопок в набор
            keyboard_answer.add(button_no)   
            
            wb.save(all_base)                    
            
            stroka=get_stroka_po_ankete_all(anketa)
            network_type_buyer = sheet["B"+stroka].value
            name_buyer = sheet["C"+stroka].value
            subscribers_buyer = sheet["D"+stroka].value
            statystics_buyer = sheet["E"+stroka].value
            info_buyer = sheet["F"+stroka].value
            category_ad_buyer = sheet["G"+stroka].value  
            price_buyer=sheet["H"+stroka].value
            
            number_buyer=sheet["M"+stroka].value
            
            edit_buyer = message.text
            if test_int(number_buyer):
                if int(number_buyer)>=1 and int(number_buyer)<=6:    
                    if number_buyer == '1':
                        sheet["B"+stroka].value=edit_buyer
                    if number_buyer == '2':
                        sheet["C"+stroka].value=edit_buyer
                    if number_buyer == '3':
                        sheet["D"+stroka].value=edit_buyer
                    if number_buyer == '4':
                        sheet["E"+stroka].value=edit_buyer
                    if number_buyer == '5':
                        sheet["F"+stroka].value=edit_buyer
                    if number_buyer == '6':
                        sheet["G"+stroka].value  =edit_buyer  
                    if number_buyer == '7':
                        sheet["H"+stroka].value = editing_buyer
                    wb.save(all_base)
                    bot.send_message(message.from_user.id, "Ваша анкета: " + '\n'
                                     + "🌐 1.Социальная сеть:" + sheet["B"+stroka] + '\n' 
                                     + "👨‍💻 2.Никнейм: " + sheet["C"+stroka] + '\n' 
                                     + "👥 3.Количество подписчиков: " + sheet["D"+stroka] + '\n' 
                                     + "📊 4.Статистика профиля: " + sheet["E"+stroka] + '\n' 
                                     + "🗺 5.Инфомация об аудитории: " + sheet["F"+stroka] + '\n' 
                                     + "📄 6.Категория рекламы: " + sheet["G"+stroka] + '\n' 
                                     + "💰 7. Максимальная стоимость рекламы: "+ sheet["H"+stroka] + '\n' 
                                     + "Прошу обратить внимание, если во время проверки Вашего профиля, администрация сервера обнаружит несовпадение данных, Ваша заявка будет аннулирована." 
                                     + '\n' + "Ваша анкета составлена верно?", reply_markup=keyboard_answer)
                    #bot.register_next_step_handler(message, checking_buyer)
                else:
                    bot.send_message(message.from_user.id,"Ошибка! Введите пункт еще раз!")
                    bot.register_next_step_handler(message, number_edit_buyer)  
            else:
                bot.send_message(message.from_user.id,"Ошибка! Введите пункт еще раз!")
                bot.register_next_step_handler(message, number_edit_buyer) 
        else:
            get_start_message(message)                
            
    
    thread1 = Thread(target=bot.polling, args=())
    thread2 = Thread(target=bot_checker.polling, args=())
    thread1.start()
    thread2.start()
    thread1.join()
    thread2.join()
while True:
    try:
        main()
    except:
        main()